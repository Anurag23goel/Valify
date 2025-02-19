from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook
from datetime import datetime
import os
import firebase_admin
from firebase_admin import credentials, firestore
from flask_cors import CORS
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# Initialize Firebase
cred = credentials.Certificate("serviceAccountKey.json")  # Replace with your Firebase credentials JSON file
firebase_admin.initialize_app(cred)

# Initialize Firestore DB
db = firestore.client()

# Flask app
app = Flask(__name__)
CORS(app)

# Define template path
TEMPLATE_PATH = "./latest.xlsm"
OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

@app.route('/download-report-pdf', methods=['GET'])
def download_report_pdf():
    try:
        # Get UID and project_id from query parameters
        uid = request.args.get('uid')
        project_id = request.args.get('project_id')
        
        if not uid or not project_id:
            return jsonify({"error": "uid and project_id are required parameters"}), 400
        
        # Load the Excel file
        workbook = load_workbook(TEMPLATE_PATH, data_only=True)
        if "Report" not in workbook.sheetnames:
            return jsonify({"error": "Sheet named 'Report' not found"}), 404
        
        sheet = workbook["Report"]
        
        # Prepare PDF file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_filename = f"report_{timestamp}.pdf"
        pdf_path = os.path.join(OUTPUT_DIR, pdf_filename)
        
        c = canvas.Canvas(pdf_path, pagesize=letter)
        width, height = letter
        
        y_position = height - 40
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
            c.drawString(40, y_position, row_text)
            y_position -= 20
            if y_position < 40:
                c.showPage()
                y_position = height - 40
        
        c.save()
        
        return send_file(pdf_path, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
